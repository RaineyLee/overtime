import os
import sys
# import warnings
# import time

import openpyxl
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5 import uic
from datetime import datetime
from openpyxl.styles import Alignment

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소
dept_main_window= uic.loadUiType(resource_path("C:\\myproject\\python project\\overtime\\overtime_v1.1\\ui\\dept_ref.ui"))[0] # Window 사용시 ui 주소
dept_window = uic.loadUiType(resource_path("C:\\myproject\\python project\\overtime\\overtime_v1.1\\ui\\dept_window.ui"))[0]
emp_window = uic.loadUiType(resource_path("C:\\myproject\\python project\\overtime\\overtime_v1.1\\ui\\emp_window.ui"))[0]

# dial_window= uic.loadUiType(resource_path("C:\\myproject\\python project\\overtime\\popup_dept_info.ui"))[0] # Window 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class DeptMainWindow(QWidget, dept_main_window) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("부서별 잔업시간 조회")
        self.slots()
        # self.date_edit.setDate(QDate.currentDate())
        # self.date = self.date_edit.date().toString("yyyyMMdd")

    def slots(self):
        self.btn_search.clicked.connect(self.make_data)
        self.btn_close.clicked.connect(self.window_close)
        self.btn_select_dept.clicked.connect(self.popup_dept_info)
        self.btn_select_emp.clicked.connect(self.popup_emp_info)
        self.btn_clear.clicked.connect(self.clear)
        self.btn_download.clicked.connect(self.make_file)

    # def set_date(self):
    #     date = self.date_select.date()
    #     self.txt_date.setText(date.toString("yyyy-MM"))

    def clear(self):        
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        self.txt_dept_id.setText("")
        self.txt_dept_name.setText("")
        self.txt_emp_id.setText("")
        self.txt_emp_name.setText("")

    def make_data(self):
        dept_id = self.txt_dept_id.toPlainText()
        emp_id = self.txt_emp_id.toPlainText()

        date_1 = self.from_date.date()
        date_2 = self.to_date.date()

        from_date = date_1.toString("yyyy-MM")
        to_date = date_2.toString("yyyy-MM")

        if  dept_id == "":
            self.msg_box("입력누락", "부서를 선택하세요")
        elif dept_id and emp_id == "":            
            dept_id = self.txt_dept_id.toPlainText()
            arr = [from_date, to_date, dept_id]

            from db.db_select import Select
            select = Select()
            result = select.dept_overtime(arr)

            if result is None:
                return

            title = ["부서명", "사원명", "날짜", "잔업시간"]
            self.make_table(len(result), result, title)
        elif dept_id and emp_id:
            dept_id = self.txt_dept_id.toPlainText()
            arr = [from_date, to_date, emp_id]

            from db.db_select import Select
            select = Select()
            result = select.emp_overtime(arr)

            if result is None:
                return

            title = ["부서명", "사원명", "날짜", "잔업시간"]
            self.make_table(len(result), result, title)
        else:
            self.msg_box("입력오류", "입력값을 확인 하세요.")

    def make_table(self, num, arr_1, title):   
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(title)

        self.tbl_info.setRowCount(num)
        self.tbl_info.setColumnCount(col)
        self.tbl_info.setHorizontalHeaderLabels(title)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_info
        header = table.horizontalHeader()

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    # 테이블 선택범위 삭제
    def delete_rows(self):
        indexes = []
        rows = []

        for idx in self.tbl_info.selectedItems():
            indexes.append(idx.row())

        for value in indexes:
            if value not in rows:
                rows.append(value)

        # 삭제시 오류 방지를 위해 아래서 부터 삭제(리버스 소팅)
        rows = sorted(rows, reverse=True)

        # 선택행 삭제
        for rowid in rows:
            self.tbl_info.removeRow(rowid)

    # def upload(self):
    #     # 현재 테이블 데이터(수정, 삭제 될 수 있다.)
    #     rows = self.tbl_info.rowCount()
    #     cols = self.tbl_info.columnCount()

    #     list = [] # 최종적으로 사용할 리스트는 for문 밖에 선언
    #     for i in range(rows):
    #         list_1 = []
    #         for j in range(cols):
    #             data = self.tbl_info.item(i,j)
    #             list_1.append(data.text())
    #         list.append(list_1)

    #     from db.db_insert import Insert
    #     data_insert = Insert()
    #     result = data_insert.insert_overtime(list)

    #     self.msg_box(result[0], result[1])

    # 부서명 가져오기 팝업
    def popup_dept_info(self):
        input_dialog = DeptWindow()
        if input_dialog.exec_():
            value = input_dialog.get_input_value()

        try:
            self.txt_dept_id.setText(value[0].text())
            self.txt_dept_name.setText(value[1].text())
        except:
            return
        
    ### 다이알로그 창으로 값을 전달 할 때는 아규먼트를 보내 주는 방식으로 !!!!
    def popup_emp_info(self):
        arg_1 = self.txt_dept_id.toPlainText()
        input_dialog = EmpWindow(arg_1) ##   <-----중요 포인트
        if input_dialog.exec_():
            value = input_dialog.get_input_value()

        try:
            self.txt_emp_id.setText(value[2].text())
            self.txt_emp_name.setText(value[3].text())
        except:
            return
        
    def get_dept_id(self):
        print(self.dept_id)
        return self.dept_id
        
     # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows = self.tbl_info.rowCount()
        cols = self.tbl_info.columnCount()

        list_2 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언

        for i in range(rows):
            list_1 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols): 
                data = self.tbl_info.item(i,j)
                list_1.append(data.text())
            list_2.append(list_1)
        
        print(list_2)

        num = len(list_2)
        self.make_excel(list_2, num)
        

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, arr, num):
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='잔업정보')

        sheet = wb.active
        list_line = ["부서명", "사원명", "날짜", "잔업시간"]
        sheet.append(list_line)

        for i in range(num):
            for j in range(len(list_line)):
                sheet.cell(row=i+2, column=j+1, value=arr[i][j])

        ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
        for column_cells in sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))


    # 파일 저장 대화상자
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)


    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

        self.close()

    # # def dept_name(self, arg_1):  
    # #     self.txt_dept_id.setText("arg_1.text()")
    # #     print(arg_1)

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()                                 # 메세지박스 실행

    def window_close(self):
        self.close()


class DeptWindow(QDialog, dept_window):
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("부서선택")
        self.slots()

        self.make_table()

    def slots(self):
        ### 다이알로그 시그널 생성기 반드시!!!!!!!! 필요. 없으면 작동 안 함############
        self.btn_confirm.clicked.connect(self.accept) # Close the dialog when OK is clicked 

    def make_table(self):
        from db.db_select import Select
        select = Select()
        select_dept = select.select_department()

        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        if select_dept is None:
            num = 0
        else:
            num = len(select_dept)
        col = self.tbl_info.columnCount()

        self.tbl_info.setRowCount(num)
        self.tbl_info.setColumnCount(col)
        # self.tbl_info.setHorizontalHeaderLabels(column_title)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_info.setItem(i, j, QTableWidgetItem(select_dept[i][j]))

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_info
        # header = table.horizontalHeader()
        table.setColumnWidth(0, int(table.width() * 0.5))
        table.setColumnWidth(1, int(table.width() * 0.5))

        # for i in range(col):
        #     header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def get_input_value(self):
        list = self.tbl_info.selectedItems()
        return list
    
    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()                                 # 메세지박스 실행

    def window_close(self):
        self.close()

class EmpWindow(QDialog, emp_window):
    def __init__(self, arg_1) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("사원선택")
        self.slots()

        self.dept_id = arg_1

        self.make_table()

    def slots(self):
        ### 다이알로그 시그널 생성기 반드시!!!!!!!! 필요. 없으면 작동 안 함############
        self.btn_confirm.clicked.connect(self.accept) # Close the dialog when OK is clicked 

    def make_table(self):
        from db.db_select import Select
        select = Select()
        select_emp = select.select_employee(self.dept_id)

        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        if select_emp is None:
            num = 0
        else:
            num = len(select_emp)
        col = self.tbl_info.columnCount()

        self.tbl_info.setRowCount(num)
        self.tbl_info.setColumnCount(col)
        # self.tbl_info.setHorizontalHeaderLabels(column_title)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_info.setItem(i, j, QTableWidgetItem(select_emp[i][j]))

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_info
        # header = table.horizontalHeader()
        table.setColumnWidth(0, int(table.width() * 0.25))
        table.setColumnWidth(1, int(table.width() * 0.25))
        table.setColumnWidth(2, int(table.width() * 0.25))
        table.setColumnWidth(3, int(table.width() * 0.25))

        # for i in range(col):
        #     header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def get_input_value(self):
        list = self.tbl_info.selectedItems()
        return list

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()                                 # 메세지박스 실행

    def window_close(self):
        self.close()

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = DeptMainWindow()
    myWindow.show()
    app.exec_()