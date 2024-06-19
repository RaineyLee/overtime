import os
import sys
# import warnings
# import time

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QDate, QSize
from PyQt5 import uic
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
total_overtime= uic.loadUiType(resource_path("./ui/total_overtime.ui"))[0] # Window 사용시 ui 주소
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class MainWindow(QWidget, total_overtime) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("잔업시간 조회")
        self.slots()
        self.date_select_1.setDate(QDate.currentDate())
        self.date_select_2.setDate(QDate.currentDate())
        self.setFixedSize(QSize(1286,817))

        self.txt_dept_id.setAlignment(Qt.AlignRight)
        self.txt_dept_id.setAlignment(Qt.AlignCenter)
        
    def slots(self):
        self.btn_search.clicked.connect(self.make_data)
        self.btn_search_dept.clicked.connect(self.popup_dept_info)
        self.btn_clear.clicked.connect(self.clear)
        self.btn_close.clicked.connect(self.close)
        self.btn_download.clicked.connect(self.make_file)
        # self.btn_close.clicked.connect(self.window_close)
        # self.btn_select_emp.clicked.connect(self.popup_emp_info)

    # def set_date(self):
    #     date = self.date_select.date()
    #     self.txt_date.setText(date.toString("yyyy-MM"))

    def clear(self):        
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        self.txt_dept_id.setText("")
        self.txt_dept_name.setText("")

    def make_data(self):
        dept_id = self.txt_dept_id.toPlainText()
        date_1 = self.date_select_1.date().toString("yyyy-MM-dd")
        date_2 = self.date_select_2.date().toString("yyyy-MM-dd")

        if  dept_id:
            arr_1 = [dept_id, date_1, date_2]
           
            from db.db_select import Select
            select = Select()

            result = select.all_overtime_1(arr_1)
            if result is None:
                return            
            else:
                title = ["부서아이디", "부서명", "사번", "이름", "날짜", "잔업시간", "시작시간", "종료시간", "작업내용", "비고"]
                self.make_table(len(result), result, title)
        elif dept_id == "":
            arr_1 = [date_1, date_2]
           
            from db.db_select import Select
            select = Select()

            result = select.all_overtime_2(arr_1)
            if result is None:
                return
            else:
                title = ["부서아이디", "부서명", "사번", "이름", "날짜", "잔업시간", "시작시간", "종료시간", "작업내용", "비고"]
                self.make_table(len(result), result, title)

    def make_table(self, num, arr_1, title):   
        self.tbl_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(title)

        self.tbl_info.setRowCount(num)
        self.tbl_info.setColumnCount(col)
        self.tbl_info.setHorizontalHeaderLabels(title)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_info
        header = table.horizontalHeader()

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    # 테이블 선택범위 삭제
    def delete_rows(self):
        indexes = []
        rows = []

        for idx in self.tbl_info.selectedItems():
            indexes.append(idx.row())

        for value in indexes:
            if value not in rows:
                rows.append(value)

        # 삭제시 오류 방지를 위해 아래서 부터 삭제(리버스 소팅)
        rows = sorted(rows, reverse=True)

        # 선택행 삭제
        for rowid in rows:
            self.tbl_info.removeRow(rowid)

    # 부서명 가져오기 팝업
    def popup_dept_info(self):
        from popup.dept_popup import DeptWindow
        input_dialog = DeptWindow()

        if input_dialog.exec_():
            value = input_dialog.get_input_value()

            try:
                self.txt_dept_id.setText(value[0].text())
                self.txt_dept_name.setText(value[1].text())
            except:
                return
        
    ### 다이알로그 창으로 값을 전달 할 때는 아규먼트를 보내 주는 방식으로 !!!!
    # def popup_emp_info(self):
    #     arg_1 = self.txt_dept_id.toPlainText()
    #     input_dialog = EmpWindow(arg_1) ##   <-----중요 포인트
    #     if input_dialog.exec_():
    #         value = input_dialog.get_input_value()

    #     try:
    #         self.txt_emp_id.setText(value[2].text())
    #         self.txt_emp_name.setText(value[3].text())
    #     except:
    #         return
        
    def get_dept_id(self):
        print(self.dept_id)
        return self.dept_id
    
      # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows = self.tbl_info.rowCount()
        cols = self.tbl_info.columnCount()

        list_2 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언

        for i in range(rows):
            list_1 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols): 
                data = self.tbl_info.item(i,j)
                list_1.append(data.text())
            list_2.append(list_1)

        num = len(list_2)
        self.make_excel(list_2, num)
        

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, arr, num):
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='잔업정보')

        sheet = wb.active
        list_line = ["부서아이디", "부서명", "사번", "이름", "날짜", "잔업시간", "시작시간", "종료시간", "작업내용", "비고"]
        sheet.append(list_line)

        for i in range(num):
            for j in range(len(list_line)):
                sheet.cell(row=i+2, column=j+1, value=arr[i][j])

        ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
        for column_cells in sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))


    # 파일 저장 대화상자(파일명 만들기)
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)

    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()                                 # 메세지박스 실행

    def window_close(self):
        self.close()

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = MainWindow()
    myWindow.show()
    app.exec_()