import os
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QSize, Qt
from PyQt5 import uic, QtWidgets
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
# 차트 생성용
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# matplotlib 폰트 설정
plt.rcParams['font.family'] ='Malgun Gothic'
plt.rcParams['axes.unicode_minus'] =False

#UI파일 연결
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소
main_window= uic.loadUiType(resource_path("./ui/main_window.ui"))[0] # Window 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, main_window) :
    def __init__(self) :
        super().__init__()

        self.version = 1.5

        from db.db_select import Select
        select = Select()
        result = select.select_version()

        if self.version == result[0][0]:
            self.setupUi(self)
            self.setWindowTitle("DOOCH PUMP HR")
            self.setFixedSize(QSize(1337,839)) # 해상도에 따라 구성 비율이 변경되게 하고 싶지 않은 경우 창의 크기를 고정 시킨다.
            self.check_login()
        else:
            self.msg_box("확인", "사용중인 프로그램의 버전 확인이 필요합니다.")
            return
        
        self.slots()

        #  # Initialize chart layout
        # self.chart_widget = QWidget(self)
        # self.layout_chart = QVBoxLayout(self.chart_widget)
        # self.setCentralWidget(self.chart_widget)

        #  차트 그리기를 위한 레이아웃, 캔버스... 초기화 
        #  attribute를 찾을 수 없다는 에러 혹은 경고 메시지가 보일 때 
        #  init에서 선언 해야 한다.
        # self.canvas_bar = None
        # self.canvas_pie = None

        # self.layout_chart = QVBoxLayout()
        # self.layout

    def slots(self):
        self.btn_refresh.clicked.connect(self.refresh_report)
        self.btn_download.clicked.connect(self.make_file)

    def check_login(self): 
        text, ok = QInputDialog.getText(self, 'Input Dialog', '사용자 번호 :')
        if ok:
            id = self.version
            password = text
        else:
            self.setFixedSize(QSize(0,0))
            return

        from db.db_select import Select
        select = Select()
        result = select.select_password(id)

        if password == result[0]:
            self.btn_refresh.show()
            self.btn_download.show()

            self.mainwindow()
        else:
            self.msg_box("오류", "사용자 코드를 확인 하세요.")
            self.setFixedSize(QSize(0,0))
            return


    def mainwindow(self):
        menu_bar = self.menuBar()
        hr_menu = menu_bar.addMenu("인사정보")
        overtime_info = menu_bar.addMenu("잔업시간 조회")
        overtime_upload = menu_bar.addMenu("잔업시간 입력")
        
        select_all = QAction('전체 조회', self)
        select_all.setStatusTip("전체 조회")
        select_all.triggered.connect(self.select_all)

        select_dept = QAction('부서별 조회', self)
        select_dept.setStatusTip("부서별 조회")
        select_dept.triggered.connect(self.select_dept)

        select_emp = QAction('사원별 조회', self)
        select_emp.setStatusTip("사원별 조회")
        select_emp.triggered.connect(self.select_emp)

        update_emp = QAction('잔업시간 수정', self)
        update_emp.setStatusTip("잔업시간 수정")
        update_emp.triggered.connect(self.update_emp)
        
        input_emp = QAction('잔업시간 입력', self)
        input_emp.setStatusTip("잔업시간 입력")
        input_emp.triggered.connect(self.input_emp)

        upload_overtime = QAction('잔업시간 업로드', self)
        upload_overtime.setStatusTip("잔업시간 업로드")
        upload_overtime.triggered.connect(self.upload_overtime)

        emp_master = QAction('인사정보', self)
        emp_master.setStatusTip("인사정보")
        emp_master.triggered.connect(self.emp_master)

        overtime_info.addAction(select_all)
        overtime_info.addAction(select_dept)
        overtime_info.addAction(select_emp)

        overtime_upload.addAction(update_emp)
        overtime_upload.addAction(input_emp)
        overtime_upload.addAction(upload_overtime)

        hr_menu.addAction(emp_master)

        status_bar = self.statusBar()
        self.setStatusBar(status_bar)

        # # 차트 그리기를 위한 레이아웃, 캔버스... 초기화 
        # # attribute를 찾을 수 없다는 에러 혹은 경고 메시지가 보일 때 
        # # init에서 선언 하던지 자동으로 선언이 가능하게 해야 한다.
        self.canvas_bar = None
        self.canvas_pie = None

        self.monthly_dept_report()
        self.monthly_sum_report()

    def make_chart(self, column_name, result):
        # Check and remove existing canvas if it exists
        if self.canvas_bar:
            self.layout_bar.removeWidget(self.canvas_bar)
            self.canvas_bar.deleteLater()

        # Create a new figure and canvas
        fig_bar = plt.Figure()
        self.canvas_bar = FigureCanvas(fig_bar)
        self.layout_bar.addWidget(self.canvas_bar)

        # Extracting data
        year_month = column_name[1:13]  # Assuming column_name includes a '날짜' column
        overtime = result[0][1:13]      # Assuming result is a list of lists with the overtime data

        # Plotting the data
        self.ax_bar = fig_bar.add_subplot(111)
        self.bars = self.ax_bar.bar(year_month, overtime)
        self.ax_bar.set_title('월별 잔업시간')
        self.ax_bar.set_xlabel('월')
        self.ax_bar.set_ylabel('잔업시간')

        # Redraw the canvas
        self.canvas_bar.draw()

        self.canvas_bar.mpl_connect('button_press_event', self.on_click)

    def on_click(self, event):
        if event.inaxes == self.ax_bar and self.bars is not None:
            for bar in self.bars:
                if bar.contains(event)[0]:
                    # label = bar.get_x() + bar.get_width() / 2
                    # value = bar.get_height()
                    col = bar.get_x() + bar.get_width() / 2
                    month = int(col) + 1
                    
                    result = self.on_click_table_info(col)
                    label = result[0]
                    value = result[1]

                    self.show_pie_chart(label, value, month)
                    break
    
    def on_click_table_info(self, arg):
        row = self.tbl_dept_info.rowCount()
        col = int(arg) + 1

        list_value = [] 
        for i in range(row):
            value = self.tbl_dept_info.item(i,col)
            list_value.append(value.text())
        list_value = list(map(float, list_value))

        list_dept = [] 
        for i in range(row):
            dept = self.tbl_dept_info.item(i,0)
            list_dept.append(dept.text())

        return list_dept, list_value

    def show_pie_chart(self, label, value, col):
        if self.canvas_pie:
            self.layout_pie.removeWidget(self.canvas_pie)
            self.canvas_pie.deleteLater()
        
        fig_pie = plt.Figure()
        self.canvas_pie = FigureCanvas(fig_pie)
        self.layout_pie.addWidget(self.canvas_pie)

         # 파이 차트 생성
        ax_pie = fig_pie.add_subplot(111)
        # values = [value, float(100) - value]  # value와 나머지 비율 계산 (예: 100에서 value를 뺀 값)
        # labels = [label, 'Others']     # 항목과 나머지 항목의 레이블 설정
        ax_pie.pie(value, labels=label, autopct='%1.1f%%')  # 파이 차트 그리기
        ax_pie.set_title(f'{col}월 부서별 잔업시간')  # 차트 제목 설정

        self.canvas_pie.draw()  # 캔버스 갱신
        
    def refresh_report(self):
        option = QtWidgets.QMessageBox.question(self, "QMessageBox", f"잔업 정보를 새로고침 하시겠습니까?", 
                                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes)
            
        if option == QtWidgets.QMessageBox.Cancel:
            return
        elif option == QtWidgets.QMessageBox.No:
            return
        elif option == QtWidgets.QMessageBox.Yes: 
            self.monthly_dept_report()
            self.monthly_sum_report()

    def monthly_dept_report(self):
        self.lbl_dept.show()
        self.tbl_dept_info.show()
        
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_dept_monthly()

        self.make_dept_table(len(result), result, column_names)
       
    def monthly_sum_report(self):
        self.tbl_month_info.show()
        
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_monthly_sum()

        self.make_sum_table(len(result), result, column_names)
        self.make_chart(column_names, result)

    def make_dept_table(self, num, arr_1, column_names):   
        self.tbl_dept_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_dept_info.setRowCount(num)
        self.tbl_dept_info.setColumnCount(col)
        self.tbl_dept_info.setHorizontalHeaderLabels(column_names)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_dept_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_dept_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_dept_info
        header = table.horizontalHeader()

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_dept_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def make_sum_table(self, num, arr_1, column_names):   
        self.tbl_month_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_month_info.setRowCount(num)
        self.tbl_month_info.setColumnCount(col)
        # self.tbl_month_info.setHorizontalHeaderLabels(column_names) #헤더 숨기기를 위해 라벨을 설정하지 않음

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_month_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_month_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_month_info
        header = table.horizontalHeader()
        header.hide() # 헤더 숨기기 함수

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_month_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

     # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows_dept_table = self.tbl_dept_info.rowCount()
        cols_dept_table = self.tbl_dept_info.columnCount()

        list_dept_1 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언
        for i in range(rows_dept_table):
            list_dept_2 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols_dept_table): 
                data_dept = self.tbl_dept_info.item(i,j)
                list_dept_2.append(data_dept.text())
            list_dept_1.append(list_dept_2)

        num_dept = len(list_dept_1)

        self.make_excel(list_dept_1, num_dept)

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, list_dept_1, num_dept):
        self.msg_box("자료저장", "부서 잔업정보가 생성 됩니다.")

        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='부서잔업정보')

        dept_sheet = wb['부서잔업정보']

        column_count = self.tbl_dept_info.columnCount()
        dept_headers = []
        for col in range(column_count):
            header_item = self.tbl_dept_info.horizontalHeaderItem(col)
            if header_item:
                dept_headers.append(header_item.text())

        dept_sheet.append(dept_headers)

        for i in range(num_dept):
            for j in range(len(dept_headers)):
                dept_sheet.cell(row=i+2, column=j+1, value=list_dept_1[i][j])
        
        ## 각 칼럼에 대해서 모든 셀값의 문자열 개수에서 1.1만큼 곱한 것들 중 최대값을 계산한다.
        for column_cells in dept_sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            dept_sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in dept_sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))

    # 파일 저장 대화상자(파일명 만들기)
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)

    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

    def select_all(self):
        import total_overtime as total_overtime_window

        self.total_window = total_overtime_window.MainWindow()
        self.total_window.show()

    def select_dept(self):
        import dept_overtime as select_dept_window

        self.dept_window = select_dept_window.DeptMainWindow()
        self.dept_window.show()
    
    def select_emp(self):
        import emp_overtime as select_emp_window

        self.emp_window = select_emp_window.MainWindow()
        self.emp_window.show() 
    
    def update_emp(self):
        import emp_overtime_update as update_emp_window

        self.emp_update_window = update_emp_window.MainWindow()
        self.emp_update_window.show() 

    def input_emp(self):
        import emp_overtime_input as input_emp_window

        self.emp_input_window = input_emp_window.MainWindow()
        self.emp_input_window.show() 

    def upload_overtime(self):
        import upload as upload_window

        self.upload_window = upload_window.MainWindow()
        self.upload_window.show()

    def emp_master(self):
        import emp_info as emp_info

        self.emp_master = emp_info.MainWindow()
        self.emp_master.show()

    def window_close(self):
        self.close()

    # def upload_location(self):        
    #     import upload_location as inv_loc

    #     self.location = inv_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.location.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def upload_barcode(self):
    #     import upload_barcode as bar_loc

    #     self.barcode = bar_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.barcode.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def upload_saleslist(self):
    #     import upload_saleslist as saleslist

    #     self.saleslist = saleslist.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.saleslist.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def item_location(self):
    #     import toexcel_location as item_loc

    #     self.item_loc = item_loc.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #     self.item_loc.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    # def make_cjnumber(self):
    #     import CJ_number_v1_2 as cj_number

    #     self.cj_number = cj_number.WindowClass() #메인창에서 띄우려면 메인창을 뜻하는 self 추가
    #  self.cj_number.show() #메인창에서 띄우려면 메인창을 뜻하는 self 추가

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()       

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    try:
        myWindow = WindowClass()
        myWindow.show()
        app.exec_()
    except Exception as e:
        msg = QMessageBox()
        msg.setWindowTitle("Error")               # 제목설정
        msg.setText(str(e))                          # 내용설정
        msg.exec_()  