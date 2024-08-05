import os
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QSize, Qt
from PyQt5 import uic
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# 절대경로를 상대경로로 변경 하는 함수
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

#UI파일 연결
# main_window= uic.loadUiType(resource_path("/Users/black/projects/make_erp/main_window.ui"))[0] # Mac 사용시 ui 주소
main_window= uic.loadUiType(resource_path("./ui/emp_overtime_month.ui"))[0] # Window 사용시 ui 주소

#화면을 띄우는데 사용되는 Class 선언
class MainWindow(QWidget, main_window):
    def __init__(self):
        super().__init__()

        self.setupUi(self)
        self.setWindowTitle("월별/사원별 잔업시간")
        self.setFixedSize(QSize(1286,790))

        self.monthly_emp_report()
        self.monthly_sum_report()

        self.slots()

    def slots(self):
        self.btn_send.clicked.connect(self.monthly_emp_report)
        self.btn_download.clicked.connect(self.make_file)
        self.btn_close.clicked.connect(self.close)

    def monthly_emp_report(self):
        self.tbl_emp_info.show()
        
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_emp_monthly()

        self.make_emp_table(len(result), result, column_names)

    def make_emp_table(self, num, arr_1, column_names):   
        self.tbl_emp_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_emp_info.setRowCount(num)
        self.tbl_emp_info.setColumnCount(col)
        self.tbl_emp_info.setHorizontalHeaderLabels(column_names)

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_emp_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_emp_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_emp_info
        header = table.horizontalHeader()

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_emp_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def monthly_sum_report(self):        
        from db.db_select import Select
        select = Select()
        result, column_names = select.select_monthly_sum()

        self.make_sum_table(len(result), result, column_names)  
    
    def make_sum_table(self, num, arr_1, column_names):   
        self.tbl_emp_month_info.setRowCount(0) # clear()는 행은 그대로 내용만 삭제, 행을 "0" 호출 한다.

        col = len(column_names)

        self.tbl_emp_month_info.setRowCount(num)
        self.tbl_emp_month_info.setColumnCount(col)
        # self.tbl_month_info.setHorizontalHeaderLabels(column_names) #헤더 숨기기를 위해 라벨을 설정하지 않음

        for i in range(num):
            for j in range(col): # 아니면 10개
                self.tbl_emp_month_info.setItem(i, j, QTableWidgetItem(str(arr_1[i][j])))
                self.tbl_emp_month_info.item(i, j).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)     

        # 컨텐츠의 길이에 맞추어 컬럼의 길이를 자동으로 조절
        ################################################################
        table = self.tbl_emp_month_info
        header = table.horizontalHeader()
        header.hide() # 헤더 숨기기 함수

        for i in range(col):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        ################################################################

        # 테이블의 길이에 맞추어 컬럼 길이를 균등하게 확장
        self.tbl_emp_month_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

     # 테이블에 남겨진 정보를 엑셀로 변환
    def make_file(self):
        rows_emp_table = self.tbl_emp_info.rowCount()
        cols_emp_table = self.tbl_emp_info.columnCount()

        list_emp_1 = [] # 최종적으로 사용할 리스트는 for문 밖에 선언
        for i in range(rows_emp_table):
            list_emp_2 = [] # 2번째 for문 안쪽에서 사용할 리스트 선언
            for j in range(cols_emp_table): 
                data_emp = self.tbl_emp_info.item(i,j)
                list_emp_2.append(data_emp.text())
            list_emp_1.append(list_emp_2)

        num_emp = len(list_emp_1)

        self.make_excel(list_emp_1, num_emp)

    # 엑셀 파일을 만들고 넘겨진 배열 정보를 이용하여 sheet에 정보를 기입/저장 함.
    def make_excel(self, list_emp_1, num_emp):
        self.msg_box("자료저장", "월별/사원별 잔업정보 sheet가 생성 됩니다.")

        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title='월별,사원별 잔업정보')

        emp_sheet = wb['월별,사원별 잔업정보']

        column_count = self.tbl_emp_info.columnCount()
        emp_headers = []
        for col in range(column_count):
            header_item = self.tbl_emp_info.horizontalHeaderItem(col)
            if header_item:
                emp_headers.append(header_item.text())

        emp_sheet.append(emp_headers)

        for i in range(num_emp):
            for j in range(len(emp_headers)):
                emp_sheet.cell(row=i+2, column=j+1, value=list_emp_1[i][j])
        
        for column_cells in emp_sheet.columns:
            # length = max(len(str(cell.value))*1.1 for cell in column_cells)
            emp_sheet.column_dimensions[column_cells[0].column_letter].width = 20
            ## 셀 가운데 정렬
            for cell in emp_sheet[column_cells[0].column_letter]:
                cell.alignment = Alignment(horizontal='center')
        
        fname = self.file_save()

        try:
            if fname:
                self.save_excel(wb, fname)
        except Exception as e:
            self.msg_box("Error", str(e))

    # 파일 저장 대화상자(파일명 만들기)
    def file_save(self):
        now = datetime.now()
        arg_1 = now.strftime('%Y-%m-%d %H-%M-%S')
        adress = "./excel/download_" + arg_1 + ".xlsx"

        dialog = QFileDialog(self)
        qurl  = dialog.getSaveFileName(parent=self, caption='Save file', directory=adress)
        
        url = qurl[0]
        try:
            return url
        except Exception as e:
            QMessageBox.about(self, 'Warning', e)

    def save_excel(self, workbook, file_name):
        workbook.save(file_name)

    def msg_box(self, arg_1, arg_2):
        msg = QMessageBox()
        msg.setWindowTitle(arg_1)               # 제목설정
        msg.setText(arg_2)                          # 내용설정
        msg.exec_()       

if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = MainWindow()
    myWindow.show()
    app.exec_()
    # try:
    #     myWindow = WindowClass()
    #     myWindow.show()
    #     app.exec_()
    # except Exception as e:
    #     msg = QMessageBox()
    #     msg.setWindowTitle("Error")               # 제목설정
    #     msg.setText(str(e))                          # 내용설정
    #     msg.exec_()  